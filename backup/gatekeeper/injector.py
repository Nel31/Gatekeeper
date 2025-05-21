# gatekeeper/injector.py
from openpyxl import load_workbook
import re
from unidecode import unidecode

FIELD_TO_HEADER = {
    'code_utilisateur':           'Code/identifiant utilisateur',
    'user_full_name':             'Nom et Prénom utilisateur',
    'user_profile':               'Profil utilisateur',
    'department':                 'Direction',
    'recommendation':             'Recommandation',
    'certificateur':              'Certificateur',
    'decision':                   'Décision',
    'execution_reco_decision':    'Exécution Reco/Décision',
    'comment_review':             'Commentaire revue',
    'comment_certificateur':      'Commentaire certificateur',
    'executed_by':                'Exécuté par',
    'execution_comment':          'Commentaire exécution'
}

def normalize(text: str) -> str:
    return re.sub(r'[^a-z0-9]', '', unidecode(text or '').lower())

def inject_to_template(report, template_path: str, output_path: str):
    wb = load_workbook(template_path)
    ws = wb.active
    header_map = {normalize(c.value): c.column for c in ws[1]}
    for i, row in enumerate(report.itertuples(index=False), start=2):
        for field, val in zip(report.columns, row):
            col = header_map.get(normalize(FIELD_TO_HEADER[field]))
            if col:
                ws.cell(row=i, column=col, value=val)
    wb.save(output_path)
