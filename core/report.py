from openpyxl import load_workbook
from datetime import datetime
from core.text_utils import normalize_text
import pandas as pd
import os

FIELD_TO_HEADER = {
    'code_utilisateur':           'Code/identifiant utilisateur',
    'nom_prenom':                 'Nom et Prénom utilisateur',
    'profil':                     'Profil utilisateur',
    'direction':                  'Direction',
    'recommendation':             'Recommandation',
    'certificateur':              'Certificateur',
    'decision':                   'Décision',
    'execution_reco_decision':    'Exécution Reco/Décision',
    'comment_review':             'Commentaire revue',
    'comment_certificateur':      'Commentaire certificateur',
    'executed_by':                'Exécuté par',
    'execution_comment':          'Commentaire exécution',
    'anomalie':                   'Anomalie',
    'date_certification':         'Date certification'
}

def normalize(text: str) -> str:
    """Normaliser un texte pour la comparaison"""
    return normalize_text(text, remove_stop_words=False)

def inject_to_template(report_df, template_path: str, output_path: str, certificateur: str = ""):
    wb = load_workbook(template_path)
    ws = wb.active
    header_map = {normalize(cell.value): cell.column for cell in ws[1] if cell.value}
    date_certif = datetime.now().strftime('%Y-%m-%d')
    
    for i, row in enumerate(report_df.itertuples(index=False), start=2):
        for field, val in zip(report_df.columns, row):
            # TOUJOURS utiliser valeurs RH pour profil/direction
            if field == 'profil':
                val = getattr(row, 'profil_rh', val) or val
            elif field == 'direction':
                val = getattr(row, 'direction_rh', val) or val
                
            template_header = FIELD_TO_HEADER.get(field)
            if not template_header:
                continue
                
            col = header_map.get(normalize(template_header))
            if col:
                if field == 'date_certification':
                    ws.cell(row=i, column=col, value=date_certif)
                else:
                    ws.cell(row=i, column=col, value=val)
    
    wb.save(output_path)
    print(f"✅ Rapport généré avec valeurs RH dans {output_path}")

def generer_rapport(df, output_path):
    """
    Génère un rapport Excel détaillé des anomalies et décisions.
    
    Args:
        df (pd.DataFrame): DataFrame contenant les données à exporter
        output_path (str): Chemin du fichier Excel de sortie
    """
    # Créer un writer Excel
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    
    # Préparer les données pour chaque feuille
    # 1. Résumé
    summary = pd.DataFrame({
        'Métrique': [
            'Total des cas',
            'Cas automatiques',
            'Cas à vérifier',
            'Variations',
            'Changements',
            'Décisions prises',
            'À conserver',
            'À désactiver'
        ],
        'Valeur': [
            len(df),
            len(df[df['cas_automatique']]),
            len(df[~df['cas_automatique']]),
            len(df[df['type_changement'].apply(lambda x: any('variation' in str(v).lower() for v in x.values()))]),
            len(df[df['type_changement'].apply(lambda x: any('changement' in str(v).lower() for v in x.values()))]),
            len(df[df['decision_finale'].notna()]),
            len(df[df['decision_finale'] == 'Conserver']),
            len(df[df['decision_finale'] == 'Désactiver'])
        ]
    })
    
    # 2. Tous les cas
    all_cases = df.copy()
    all_cases['Type de changement'] = all_cases['type_changement'].apply(
        lambda x: ' | '.join(f"{k}: {v}" for k, v in x.items())
    )
    
    # 3. Cas automatiques
    auto_cases = df[df['cas_automatique']].copy()
    auto_cases['Type de changement'] = auto_cases['type_changement'].apply(
        lambda x: ' | '.join(f"{k}: {v}" for k, v in x.items())
    )
    
    # 4. Cas à vérifier
    manual_cases = df[~df['cas_automatique']].copy()
    manual_cases['Type de changement'] = manual_cases['type_changement'].apply(
        lambda x: ' | '.join(f"{k}: {v}" for k, v in x.items())
    )
    
    # 5. Variations
    variations = df[df['type_changement'].apply(
        lambda x: any('variation' in str(v).lower() for v in x.values())
    )].copy()
    variations['Type de changement'] = variations['type_changement'].apply(
        lambda x: ' | '.join(f"{k}: {v}" for k, v in x.items())
    )
    
    # 6. Changements
    changements = df[df['type_changement'].apply(
        lambda x: any('changement' in str(v).lower() for v in x.values())
    )].copy()
    changements['Type de changement'] = changements['type_changement'].apply(
        lambda x: ' | '.join(f"{k}: {v}" for k, v in x.items())
    )
    
    # Écrire chaque feuille
    summary.to_excel(writer, sheet_name='Résumé', index=False)
    all_cases.to_excel(writer, sheet_name='Tous les cas', index=False)
    auto_cases.to_excel(writer, sheet_name='Cas automatiques', index=False)
    manual_cases.to_excel(writer, sheet_name='Cas à vérifier', index=False)
    variations.to_excel(writer, sheet_name='Variations', index=False)
    changements.to_excel(writer, sheet_name='Changements', index=False)
    
    # Ajuster la largeur des colonnes
    for sheet_name in writer.sheets:
        worksheet = writer.sheets[sheet_name]
        for idx, col in enumerate(worksheet.columns):
            max_length = max(
                len(str(cell.value)) for cell in col
            )
            worksheet.column_dimensions[chr(65 + idx)].width = min(max_length + 2, 50)
    
    # Sauvegarder le fichier
    writer.close()
