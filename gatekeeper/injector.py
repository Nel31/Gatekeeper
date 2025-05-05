# gatekeeper/injector.py
from openpyxl import load_workbook
import re
from unidecode import unidecode

def normalize(text: str) -> str:
    h = unidecode(text or "")
    return re.sub(r'[^a-z0-9]', '', h.lower())

FIELD_TO_HEADER = {
    'code_utilisateur':'Code/identifiant utilisateur',
    'nom_prenom':'Nom et Prénom utilisateur',
    'profil':'Profil utilisateur',
    'direction':'Direction',
    'recommandation':'Recommandation',
    'certificateur':'Certificateur',
    'commentaire_revue':'Commentaire revue',
    'décision':'Décision',
    'commentaire_certificateur':'Commentaire certificateur',
    'exécution':'Exécution Reco/Décision',
    'exécuté_par':'Exécuté par',
    'commentaire_exécution':'Commentaire exécution'
}

def inject_to_template(report, template_path: str, output_path: str):
    wb = load_workbook(template_path)
    ws = wb.active

    header_map = {
        normalize(cell.value): cell.column
        for cell in ws[1]
    }

    for i, row in enumerate(report.itertuples(index=False), start=2):
        for field, value in zip(report.columns, row):
            col = header_map[normalize(FIELD_TO_HEADER[field])]
            ws.cell(row=i, column=col, value=value)

    wb.save(output_path)
