from openpyxl import load_workbook
import pandas as pd

def write_excel(out_path: str, df: pd.DataFrame, template_path: str):
    wb = load_workbook(template_path)
    ws = wb.active
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(out_path)
